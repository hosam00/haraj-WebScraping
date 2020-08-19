from time import sleep

from bs4 import BeautifulSoup
from requests import get
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotInteractableException


def extract_urls():
    url = 'https://haraj.com.sa/tags/%D8%A7%D9%84%D8%B1%D9%8A%D8%A7%D8%B6_%D8%AE%D9%8A%D9%84'
    driver = webdriver.Chrome(executable_path="E:/test/chromedriver.exe")
    driver.maximize_window()
    driver.get(url)
    number = 1
    # links = [l.get_attribute('href') for l in driver.find_elements_by_css_selector('.postTitle a')]
    links = []
    while number <= 49:
        html = driver.find_element_by_tag_name('html')
        html.send_keys(Keys.END)
        # sleep(1)
        if driver.find_element_by_xpath('//*[@id="more"]') is not None:
            driver.find_element_by_xpath('//*[@id="more"]').click()
        links += [l.get_attribute('href') for l in driver.find_elements_by_css_selector('.postTitle a')]
        print(number)
        number += 1
    links = list(set(links))
    with open('links.txt', 'a') as file:
        for link in links:
            file.write(link)
            file.write('\n')
    driver.quit()


def extract_data():
    driver = webdriver.Chrome(executable_path="E:/test/chromedriver.exe")
    driver.maximize_window()
    wb = load_workbook('haraj.xlsx')
    sheet = wb.active
    row = 2
    with open('links.txt', 'r') as file:
        for link in file:
            link = link.strip()
            driver.get(link)
            try:
                name = driver.find_element_by_css_selector('.postHeader h3').text
            except:
                name = ""
            try:
                city = driver.find_element_by_css_selector('.postExtraInfoPart:nth-child(1) a').text
            except:
                city = ""
            try:
                publisher = driver.find_element_by_css_selector('.postExtraInfoPart+ .postExtraInfoPart a').text
            except:
                publisher = ""
            try:
                description = driver.find_element_by_css_selector('#root > div > div.postWrapper > div.postMain > div.postViewContainer > div.postBody').text
            except:
                description = ""
            try:
                phone = driver.find_element_by_css_selector('#root > div > div.postWrapper > div.postMain > div.postViewContainer > div.postContact > strong').text
            except:
                phone = ""
            try:
                date = driver.find_element_by_css_selector('#root > div > div.postWrapper > div.postMain > div.postViewContainer > div.postHeader > div:nth-child(3) > div:nth-child(1) > span').text
            except:
                date = ""
            sheet.cell(row=row, column=1, value=name)
            sheet.cell(row=row, column=2, value=city)
            sheet.cell(row=row, column=3, value=description)
            sheet.cell(row=row, column=4, value=publisher)
            sheet.cell(row=row, column=5, value=date)
            sheet.cell(row=row, column=6, value=phone)
            sheet.cell(row=row, column=7, value=link)
            print(row)
            row += 1
    wb.save('haraj.xlsx')
    driver.quit()


# extract_urls()
extract_data()
